import requests
import urllib3
from urllib.error import HTTPError
import json
from lxml import etree
# excel相关
from openpyxl import load_workbook
import warnings

# 忽视openpyxl警告
warnings.filterwarnings('ignore')
# 忽视request警告
urllib3.disable_warnings()

########################################################################################
# 自定义配置
########################################################################################
# 目标网站首页
mainUrl = 'https://script.hznmd.com/api/goods/list?page=1&sort=sale_count%7Cdesc&type=0'
# 详情属性对照表
goodsPropertyConfigUrl = 'https://script.hznmd.com/api/getGoodsPropertyConfig'
# EXCEL模版路径
excelPath = './storyModel.xlsx'
# EXCEL起始行
startRow = 3
# EXCEL起始列
startCol = 1
########################################################################################

# 全局变量:详情属性对照表
    # 2021.08.09结构
    ##### 0: script_role_num
    ##### 1: script_background
    ##### 2: script_theme
    ##### 3: script_type
    ##### 4: script_difficulty
    ##### 5: script_sell_format
goodsPropertyConfig = []
# 全局变量:剧本计数ID
storyNumId = 0


# 主流程
def main():
    # 全局变量:详情属性对照表
    global goodsPropertyConfig
    # 获取详情属性对照表
    goodsPropertyConfig = getgoodsPropertyConfig()
    # 获取总页数
    totalPageNum = getPageNum()
    # 获取单页ID
    getPageStoryId(totalPageNum)


# 获取详情属性对照表
def getgoodsPropertyConfig():
    res = getRequest(goodsPropertyConfigUrl)['data']
    return res


# 获取总页数
def getPageNum():
    resJson = getRequest(mainUrl)
    resPageNum = resJson['data']['last_page']
    return resPageNum


# 获取总剧本数
def getStoryNum():
    resJson = getRequest(mainUrl)
    resStoryNum = resJson['data']['total']
    return resStoryNum


# 获取每一页上剧本ID
def getPageStoryId(totalPageNum):
    # 全局变量:剧本计数ID
    global storyNumId
    storyNum = getStoryNum()
    for i in range(1, totalPageNum):
        # 单页url拼接
        onePageUrl = 'https://script.hznmd.com/api/goods/list?page=' + str(i) + '&sort=sale_count%7Cdesc&type=0'
        # 单页详情
        onePageJson = getRequest(onePageUrl)
        # 单页剧本信息
        onePageDetail = onePageJson['data']['data']
        # 单页剧本数量
        onePageDetailNum = len(onePageDetail)
        # 循环获取id
        for k in range(onePageDetailNum):
            storyId = onePageDetail[k]['id']
            getPageStoryDetail(storyId)
            storyNumId = storyNumId + 1
    print('爬虫结束：应获取' + str(storyNum) + '条，实际获取：' + str(storyNumId) + '条')

# 获取每一页上剧本详情
def getPageStoryDetail(storyId):
    detailUrl = 'https://script.hznmd.com/api/getGoods/' + str(storyId)
    detailJson = getRequest(detailUrl)['data']
    # 获取剧本名称
    outputName = getDetailName(detailJson)
    # 获取剧本封面
    outputThumbnail = getDetailThumbnail(detailJson)
    # 获取剧本文字详情
    outputInfoText = getDetailInfoText(detailJson)
    # 获取剧本图片详情
    outputInfoImg = getDetailInfoImg(detailJson)
    # 获取剧本主题
    outputTheme = getDetailTheme(detailJson)
    # 获取男生人数
    outputBoyNum = getBoyNum(detailJson)
    if outputBoyNum is None:
        outputBoyNum = 0
    # 获取女生人数
    outputGirlNum = getGirlNum(detailJson)
    if outputGirlNum is None:
        outputGirlNum = 0
    # 获取总人数
    outputTotalNum = outputBoyNum + outputGirlNum
    # 获取游戏时长
    outputGameHour = getGameHour(detailJson)
    # EXCEL操作
    excelMethods(storyId, outputName, outputThumbnail, outputInfoText, outputInfoImg, outputTheme, outputBoyNum, outputGirlNum, outputTotalNum, outputGameHour)


########################################################################################
# 剧本详情类方法
########################################################################################
# 获取剧本名称
def getDetailName(detailJson):
    # 剧本名称
    detailName = detailJson['name']
    detailName = detailName.replace('【', '')
    detailName = detailName.replace('】', '')
    return detailName


# 获取剧本封面
def getDetailThumbnail(detailJson):
    # 剧本封面
    thumbnail = detailJson['thumb_urls'][0]
    return thumbnail


# 获取剧本文字详情
def getDetailInfoText(detailJson):
    # 剧本文字详情
    infoText = detailJson['html_detail']
    # 为空时回调
    if infoText == '':
        return ''
    xpathSelector = etree.HTML(infoText)
    xpathRes = xpathSelector.xpath('//p/text()')
    # 拼接字符串
    string = ';'
    infoText = string.join(xpathRes)
    return infoText


# 获取剧本图片详情
def getDetailInfoImg(detailJson):
    # 剧本图片详情
    infoImg = detailJson['html_detail']
    # 为空时回调
    if infoImg == '':
        return ''
    xpathSelector = etree.HTML(infoImg)
    xpathRes = xpathSelector.xpath('//img/@src')
    # 拼接字符串
    string = ';'
    infoImg = string.join(xpathRes)
    return infoImg


# 获取剧本主题
def getDetailTheme(detailJson):
    # 剧本主题
    global goodsPropertyConfig
    # 1、背景
    background = detailJson['script_background']
    themeList = []
    for backgroundItem in background:
        for item in goodsPropertyConfig[1]['options']:
            if item['id'] == backgroundItem:
                themeList.append(item['text'])
    # 2、题材
    theme = detailJson['script_theme']
    for themeItem in theme:
        for item in goodsPropertyConfig[2]['options']:
            if item['id'] == themeItem:
                themeList.append(item['text'])
    # 3、类型
    type = detailJson['script_type']
    for typeItem in type:
        for item in goodsPropertyConfig[3]['options']:
            if item['id'] == typeItem:
                themeList.append(item['text'])
    # 4、难度
    difficulty = detailJson['script_difficulty']
    for difficultyItem in difficulty:
        for item in goodsPropertyConfig[4]['options']:
            if item['id'] == difficultyItem:
                themeList.append(item['text'])
    # 5、发售形式
    sellFormat = detailJson['script_sell_format']
    for sellFormatItem in sellFormat:
        for item in goodsPropertyConfig[5]['options']:
            if item['id'] == sellFormatItem:
                themeList.append(item['text'])
    # 拼接字符串
    string = ','
    themeRes = string.join(themeList)
    return themeRes


# 获取男生人数
def getBoyNum(detailJson):
    # 男生人数
    boyNum = detailJson['boy_num']
    return boyNum


# 获取女生人数
def getGirlNum(detailJson):
    # 女生人数
    girlNum = detailJson['girl_num']
    return girlNum


# 获取游戏时长
def getGameHour(detailJson):
    # 游戏时长
    # 游戏最短时长
    gameHourMin = detailJson['game_min_hour']
    # 游戏最长时长
    gameHourMax = detailJson['game_max_hour']
    if gameHourMin is None:
        gameHourMin = 0
    if gameHourMax is None:
        gameHourMax = 0
    gameHour = str(gameHourMin) + '-' + str(gameHourMax) + '小时'
    return gameHour
########################################################################################


########################################################################################
# EXCEL操作
########################################################################################
def excelMethods(storyId, outputName, outputThumbnail, outputInfoText, outputInfoImg, outputTheme, outputBoyNum, outputGirlNum, outputTotalNum, outputGameHour):
    # 全局变量:剧本计数ID
    global storyNumId
    # 读取EXCEL
    wb = load_workbook(excelPath)
    ws = wb.active
    # 获取剧本ID
    ws.cell(startRow + storyNumId, startCol).value = storyId
    # 获取剧本名称
    ws.cell(startRow + storyNumId, startCol + 1).value = outputName
    # 获取剧本封面
    ws.cell(startRow + storyNumId, startCol + 2).value = outputThumbnail
    # 获取剧本文字详情
    ws.cell(startRow + storyNumId, startCol + 3).value = outputInfoText
    # 获取剧本图片详情
    ws.cell(startRow + storyNumId, startCol + 4).value = outputInfoImg
    # 获取剧本主题
    ws.cell(startRow + storyNumId, startCol + 5).value = outputTheme
    # 获取男生人数
    ws.cell(startRow + storyNumId, startCol + 6).value = outputBoyNum
    # 获取女生人数
    ws.cell(startRow + storyNumId, startCol + 7).value = outputGirlNum
    # 获取总人数
    ws.cell(startRow + storyNumId, startCol + 8).value = outputTotalNum
    # 获取游戏时长
    ws.cell(startRow + storyNumId, startCol + 9).value = outputGameHour
    # 保存EXCEL
    wb.save(excelPath)
########################################################################################


########################################################################################
# 请求封装
########################################################################################
# get请求
def getRequest(url):
    try:
        response = requests.get(url, verify=False)
        # 如果状态码不是200，引发HttpError异常
        response.raise_for_status()
    except HTTPError as e:
        print(e)
    else:
        resObj = json.loads(response.text)
        print(url + ':' + str(response.status_code))
        return resObj
########################################################################################


if __name__ == '__main__':
    main()
